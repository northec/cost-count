#!/usr/bin/env python3
"""
文件统计脚本 - 统计目录中的PDF/DGN/DWG/XLS等文件并生成Excel报告
"""

import os
import sys
from pathlib import Path
from collections import defaultdict
import math

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("请先安装依赖: pip install openpyxl pypdf")
    sys.exit(1)

try:
    import pypdf
except ImportError:
    print("请先安装依赖: pip install pypdf")
    sys.exit(1)


# A4纸张尺寸（毫米）
A4_WIDTH_MM = 210
A4_HEIGHT_MM = 297
A4_AREA_MM2 = A4_WIDTH_MM * A4_HEIGHT_MM


def get_file_size(size_bytes):
    """将字节转换为可读格式（K或M）"""
    if size_bytes < 1024:
        return f"{size_bytes}B", size_bytes
    elif size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.2f}K", size_bytes / 1024
    else:
        return f"{size_bytes / (1024 * 1024):.2f}M", size_bytes / (1024 * 1024)


def get_pdf_a4_pages(file_path):
    """获取PDF文件并计算换算成A4的页数
    返回: (实际页数, A4换算页数)
    """
    try:
        with open(file_path, 'rb') as f:
            pdf_reader = pypdf.PdfReader(f)
            total_pages = len(pdf_reader.pages)
            a4_pages = 0

            for page in pdf_reader.pages:
                # 获取页面尺寸（点 point）
                # 1 inch = 72 points, 1 inch = 25.4 mm
                # 所以 1 point = 25.4 / 72 mm
                media_box = page.mediabox
                width_pt = media_box.width
                height_pt = media_box.height

                # 转换为毫米
                width_mm = width_pt * 25.4 / 72
                height_mm = height_pt * 25.4 / 72

                # 确定页面方向（取长边为高度）
                page_width = min(width_mm, height_mm)
                page_height = max(width_mm, height_mm)

                # 计算页面面积（mm²）
                page_area = page_width * page_height

                # 换算成A4页数（面积比）
                page_ratio = page_area / A4_AREA_MM2
                a4_pages += page_ratio

            return total_pages, round(a4_pages, 2)
    except Exception as e:
        print(f"读取PDF页数失败 {file_path}: {e}")
        return 0, 0


def calculate_pdf_points(a4_page_count):
    """计算PDF文件的积分（基于A4换算页数）
    - 10积分起步（含3页A4）
    - 超出3页，3积分/页
    """
    # 基础积分：10积分含3页A4
    base_points = 10
    base_pages = 3

    if a4_page_count <= base_pages:
        return base_points
    else:
        # 超出页数
        extra_pages = a4_page_count - base_pages
        extra_points = extra_pages * 3
        return base_points + extra_points


def calculate_drawing_points():
    """计算DWG/DGN文件的积分 - 按16页PDF的积分计算
    16页PDF = 10积分(前3页) + 13页*3积分 = 49积分
    """
    base_points = 10  # 前3页
    extra_pages = 16 - 3  # 13页
    return base_points + extra_pages * 3


def scan_directory(directory):
    """扫描目录获取所有文件信息"""
    file_info_list = []
    extensions = ['.pdf', '.dgn', '.dwg']

    directory = Path(directory)

    for root, dirs, files in os.walk(directory):
        for filename in files:
            file_path = Path(root) / filename
            ext = file_path.suffix.lower()

            if ext in extensions:
                file_stat = os.stat(file_path)
                size_bytes = file_stat.st_size
                size_readable, size_kb = get_file_size(size_bytes)

                info = {
                    'path': str(file_path.as_posix()),  # 使用正斜杠，跨平台兼容
                    'filename': filename,
                    'type': ext.upper().replace('.', ''),
                    'size_readable': size_readable,
                    'size_bytes': size_bytes,
                    'size_kb': size_kb,
                }

                # PDF特殊处理
                if ext == '.pdf':
                    actual_pages, a4_pages = get_pdf_a4_pages(file_path)
                    info['page_count'] = actual_pages  # 显示实际页数
                    info['a4_pages'] = a4_pages  # A4换算页数
                    info['points'] = calculate_pdf_points(a4_pages)
                # DWG/DGN特殊处理
                elif ext in ['.dwg', '.dgn']:
                    info['page_count'] = 16  # 按16页PDF计算
                    info['points'] = calculate_drawing_points()

                file_info_list.append(info)

    return file_info_list


def create_excel_report(file_info_list, output_path):
    """创建Excel报告"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "文件统计"

    # 定义样式
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    cell_alignment = Alignment(horizontal='left', vertical='center')
    number_alignment = Alignment(horizontal='right', vertical='center')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 设置列宽
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 10

    # 写入表头
    headers = ['路径', '文件名', '类型', '大小', '页数', 'A4页数', '积分']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # 写入数据
    for row_idx, info in enumerate(file_info_list, 2):
        ws.cell(row=row_idx, column=1, value=info['path']).alignment = cell_alignment
        ws.cell(row=row_idx, column=2, value=info['filename']).alignment = cell_alignment
        ws.cell(row=row_idx, column=3, value=info['type']).alignment = cell_alignment
        ws.cell(row=row_idx, column=4, value=info['size_readable']).alignment = number_alignment
        ws.cell(row=row_idx, column=5, value=info['page_count']).alignment = number_alignment
        # A4页数列
        a4_pages = info.get('a4_pages', info.get('page_count', ''))
        ws.cell(row=row_idx, column=6, value=a4_pages).alignment = number_alignment
        ws.cell(row=row_idx, column=7, value=info['points']).alignment = number_alignment

        # 添加边框
        for col in range(1, 8):
            ws.cell(row=row_idx, column=col).border = border

    # 统计数据
    total_files = len(file_info_list)
    total_size_bytes = sum(f['size_bytes'] for f in file_info_list)
    total_size_mb = total_size_bytes / (1024 * 1024)
    total_points = sum(f['points'] for f in file_info_list)

    # 按类型统计
    stats = defaultdict(lambda: {'count': 0, 'size': 0, 'pages': 0, 'points': 0})
    for f in file_info_list:
        file_type = f['type']
        stats[file_type]['count'] += 1
        stats[file_type]['size'] += f['size_bytes']
        # 统计A4换算页数
        stats[file_type]['pages'] += f.get('a4_pages', f.get('page_count', 0))
        stats[file_type]['points'] += f['points']

    # 写入统计信息
    start_row = len(file_info_list) + 3

    # 统计标题
    title_cell = ws.cell(row=start_row, column=1, value="【统计数据】")
    title_cell.font = Font(name='微软雅黑', size=12, bold=True)
    ws.merge_cells(f"{get_column_letter(1)}{start_row}:{get_column_letter(7)}{start_row}")

    # 统计表头
    stats_row = start_row + 1
    stats_headers = ['类型', '文件个数', '总体积(MB)', '总页数', '总积分']
    for col, header in enumerate(stats_headers, 1):
        cell = ws.cell(row=stats_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # 各类文件统计
    current_row = stats_row + 1
    for file_type in ['PDF', 'DGN', 'DWG']:
        if file_type in stats:
            s = stats[file_type]
            ws.cell(row=current_row, column=1, value=f'{file_type}文件').border = border
            ws.cell(row=current_row, column=2, value=s['count']).border = border
            ws.cell(row=current_row, column=2).alignment = number_alignment
            ws.cell(row=current_row, column=3, value=f"{s['size'] / (1024 * 1024):.2f}").border = border
            ws.cell(row=current_row, column=3).alignment = number_alignment
            ws.cell(row=current_row, column=4, value=s['pages'] if s['pages'] else '').border = border
            ws.cell(row=current_row, column=4).alignment = number_alignment
            ws.cell(row=current_row, column=5, value=s['points']).border = border
            ws.cell(row=current_row, column=5).alignment = number_alignment
            current_row += 1

    # 总计行
    total_row = current_row
    ws.cell(row=total_row, column=1, value='总计').font = Font(bold=True)
    ws.cell(row=total_row, column=1).border = border
    ws.cell(row=total_row, column=2, value=total_files).border = border
    ws.cell(row=total_row, column=2).alignment = number_alignment
    ws.cell(row=total_row, column=2).font = Font(bold=True)
    ws.cell(row=total_row, column=3, value=f"{total_size_mb:.2f}").border = border
    ws.cell(row=total_row, column=3).alignment = number_alignment
    ws.cell(row=total_row, column=3).font = Font(bold=True)

    # 总A4页数
    total_a4_pages = sum(s.get('pages', 0) for s in stats.values())
    ws.cell(row=total_row, column=4, value=total_a4_pages).border = border
    ws.cell(row=total_row, column=4).alignment = number_alignment
    ws.cell(row=total_row, column=4).font = Font(bold=True)

    ws.cell(row=total_row, column=5, value=total_points).border = border
    ws.cell(row=total_row, column=5).alignment = number_alignment
    ws.cell(row=total_row, column=5).font = Font(bold=True)

    # 总积分量单独显示
    summary_row = total_row + 2
    ws.cell(row=summary_row, column=1, value="总积分量")
    ws.cell(row=summary_row, column=1).font = Font(name='微软雅黑', size=12, bold=True)
    ws.cell(row=summary_row, column=2, value=total_points)
    ws.cell(row=summary_row, column=2).font = Font(name='微软雅黑', size=12, bold=True, color='FF0000')
    ws.cell(row=summary_row, column=2).alignment = number_alignment

    wb.save(output_path)
    print(f"报告已生成: {output_path}")

    # 自动打开Excel文件
    import subprocess
    import platform
    try:
        if platform.system() == 'Darwin':  # macOS
            subprocess.run(['open', output_path])
        elif platform.system() == 'Windows':
            os.startfile(output_path)
        else:  # Linux
            subprocess.run(['xdg-open', output_path])
    except Exception as e:
        print(f"无法自动打开文件: {e}")


def main():
    # 获取输入目录，默认当前目录
    if len(sys.argv) > 1:
        target_dir = sys.argv[1]
    else:
        target_dir = "."

    # 检查目录是否存在
    if not os.path.isdir(target_dir):
        print(f"错误: 目录不存在 - {target_dir}")
        sys.exit(1)

    print(f"正在扫描目录: {os.path.abspath(target_dir)}")

    # 扫描文件
    file_info_list = scan_directory(target_dir)
    print(f"找到 {len(file_info_list)} 个文件")

    if not file_info_list:
        print("没有找到支持的文件类型")
        return

    # 生成输出文件名（带时间戳）
    from datetime import datetime
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_name = f"文件统计报告_{timestamp}.xlsx"
    output_path = os.path.join(os.getcwd(), output_name)

    # 如果在参数目录中运行，保存到该目录
    if os.path.abspath(target_dir) != os.getcwd():
        output_path = os.path.join(target_dir, output_name)

    # 生成Excel报告
    create_excel_report(file_info_list, output_path)


if __name__ == "__main__":
    main()
